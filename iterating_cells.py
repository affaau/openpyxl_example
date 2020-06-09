from openpyxl import load_workbook

workbook = load_workbook(filename="hello_world.xlsx", data_only=True)

sheet = workbook.active

# Read in as tuple of tuples
# ( 
#   (<Cell 'sheet'.A1>, <Cell 'sheet'.B1>, <Cell 'sheet'.C1>)   'A1 - C1'
#   (<......>,          <....>,            <....>)              'A2 - C2'
#   (<......>,          <....>,            <Cell 'sheet'.C3>)   'A3 - C3'
# )
cells_group = sheet["A1:C3"]  # two-dimension
# other examples
#   sheet[5:6] - two rows of cells, row=5 and row=6 of EXCEL
#
# one-dimension examples
#   sheet[5] - tuple of cells of whole row=5 of EXCEL 
#   sheet["B"] - tuple of cells of whole column=2 of EXCEL

# zero-indexed
print(cells_group[0])           # cells tuple of first row
 
print(cells_group[0][0].value)  # value of first row, first column


# work with iterator
for row in sheet.iter_rows(min_row=1,
                           max_row=2,
                           min_col=1,
                           max_col=3):
    print(row)
# (<Cell 'Sheet 1'.A1>, <Cell 'Sheet 1'.B1>, <Cell 'Sheet 1'.C1>)
# (<Cell 'Sheet 1'.A2>, <Cell 'Sheet 1'.B2>, <Cell 'Sheet 1'.C2>)

###############################################################
## for iter_cols to work, read_only has to be False... WHY?! ##
###############################################################
for column in sheet.iter_cols(min_row=1,
                              max_row=2,
                              min_col=1,
                              max_col=3):
    print(column)
# (<Cell 'Sheet 1'.A1>, <Cell 'Sheet 1'.A2>)
# (<Cell 'Sheet 1'.B1>, <Cell 'Sheet 1'.B2>)
# (<Cell 'Sheet 1'.C1>, <Cell 'Sheet 1'.C2>)

# options - to show velues instead of cell objects
for column in sheet.iter_cols(min_row=1,
                              max_row=3,
                              min_col=1,
                              max_col=3, values_only=True):
    print(column)
# ("Hello", None, 1)
# ("World!", None, 2)
# (None, None, 3)

# to iterate of the whole sheet, 
# for row in sheet.rows:
#     print(row)
#
# row is tuple of whole row of cells
# (<Cell...>, <Cell...>....)


workbook.close()
