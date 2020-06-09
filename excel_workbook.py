'''Basics operation of create, write simple text to
specific cells and save
'''
from openpyxl import Workbook

workbook = Workbook()

# default sheet name is 'Sheet'
sheet = workbook.active

# change title
sheet.title = 'My first sheet'

# write to cell
sheet["A1"] = "Hello"
sheet["B1"] = "World!"

# default create as last sheet
# default name is 'Sheet' or follow EXCEL's rule
workbook.create_sheet()
workbook.create_sheet()

# index - create & insert at index (zero-based)
new_sheet = workbook.create_sheet(title="new_sheet", index=0)  # as the first sheet
# new sheet becomes the 'active' one
print(workbook.active.title)

sheet["A3"] = 1
sheet["B3"] = 2
sheet["C3"] = '=A3 + B3'      # create formula
sheet["D3"] = '=PRODUCT(A3:B3)'   # call function

# explicitly change the active sheet
workbook.active = 1    # by index
print(workbook.active.title)    # 'My first sheet'
# alternatively
workbook.active = workbook.get_sheet_by_name('My first sheet')   # by name
print(workbook.active.title)

workbook.save(filename="hello_world.xlsx")
