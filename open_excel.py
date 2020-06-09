from openpyxl import load_workbook


# default read_only=False - allow edit
#         read_only=True - advantage in reading very large files
# default data_only=False - show formula of cells
workbook = load_workbook(filename="hello_world.xlsx")

# list of all sheet names
print(workbook.sheetnames)

sheet = workbook.active
# for other sheet
# sheet = workbook[other_sheet_name]
print(sheet.title)
# sheet.title = 'new_title_name'

print(sheet["A1"].value)  # show content in cell

# another way to access cell
# in openpyxl & EXCEL, one-indexed nottation is used
print(sheet.cell(row=1, column=2).value)  # same as 'B1'

print(sheet.cell(row=3, column=3).value)  # suppose a formula

workbook.close()

workbook = load_workbook(filename="hello_world.xlsx", data_only=True)
sheet = workbook.active

print(sheet.cell(row=3, column=3).value)  # now data only!

workbook.close()
